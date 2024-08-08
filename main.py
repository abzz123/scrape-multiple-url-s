import requests
from bs4 import BeautifulSoup, NavigableString
from openpyxl import Workbook


# Function to scrape the title, description, and SKU from the given URL
def scrape_website(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    # Extracting title
    #change class
    title = soup.find('h1', class_='page-title')
    if title:
        title = title.text.strip()
    else:
        title = "Title Not Found"

    # Extracting description
    #change class
    description = soup.find('div', class_='description')
    if description:
        description = description.text.strip()
    else:
        description = "Description Not Found"
    # change class or take this section out
    description1 = soup.find('div', class_='product attribute overview')
    if description1:
        description1 = description1.text.strip()
    else:
        description1 = "Description Not Found"

    # change class or take this section out
    description2 = soup.find('div', class_='application-list')
    if description2:
        description2 = description2.text.strip()
    else:
       description2 = "Description Not Found"
#desription which will be places in the excel sheet as 1
    #change the company name and the car brand name. where it says partsplugform manually in the excel sheet click replace and manually past the actual form
    full_description = f"<h2>HardRace - VOLKSWAGEN - {title}</h2>\nspacee\n{description}\nspacee\n{description1}\nspacee\n{description2}\nspacee\n\n\n partsplugform"
    # Extracting SKU
    sku = soup.find('div', class_='product attribute sku')
    if sku:
        sku = sku.text.strip()
    else:
        sku = "SKU Not Found"

        # Extracting price
    price = soup.find('span', class_='price')
    if price:
        price = price.text.strip()
    else :
        price = "Price Not Found"

    return title, full_description, sku, price


# Function to write data into Excel spreadsheet
def write_to_excel(title, full_description, sku, price, row, ws):
    # Writing data
    #change the company name and the car brand
    title_with_sku = f'HardRace - VOLKSWAGEN - {title} - {sku}'
    ws.cell(row=row, column=1, value=title_with_sku)
    ws.cell(row=row, column=2, value=full_description)
    ws.cell(row=row, column=3, value=sku)
    ws.cell(row=row, column=4, value=price)

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Writing headers
ws['A1'] = 'Title'
ws['B1'] = 'Description'
ws['C1'] = 'SKU'
ws['D1'] = 'Price'

# URLs of the product pages as many as u like. examples below
urls = [


    'https://www.hardrace.co.uk/front-lower-arm-bush-front-0070284.html',
    'https://www.hardrace.co.uk/front-lower-arm-rear-bush-0070285.html',


]

# Scrape each website and write data to Excel
for i, url in enumerate(urls, start=2):
    title, full_description, sku, price = scrape_website(url)
    write_to_excel(title, full_description, sku, price, i, ws)

# Save the workbook
wb.save('Multiple_Urls_Products.xlsx')

print("Data has been successfully written to Multiple_Urls_Products.xlsx")